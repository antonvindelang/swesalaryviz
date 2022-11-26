from openpyxl import load_workbook
import re
import pandas as pd
import numpy as np

# Download from SCB https://www.scb.se/hitta-statistik/statistik-efter-amne/hushallens-ekonomi/inkomster-och-inkomstfordelning/inkomster-och-skatter/
WB_FILE = 'data/csfvi---inkomstklasser-2020.xlsx'
COLUMNS = ['age_range', 'gender', 'income',
           'min_income', 'max_income', 'number_of_people']
OUT_FILE_INCOME = 'data/income_classes_processed.csv'
OUT_FILE_PERCENTILES = 'data/percentiles_processed.csv'


def process_workbook(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    age_range_cells = ['C6', 'G6', 'K6', 'O6',
                       'S6', 'W6', 'AA6', 'AE6', 'AI6', 'AM6']
    age_ranges = [sheet[cell].value.replace(
        '–', '-') for cell in age_range_cells]

    gender_classes_cells = ['C7', 'D7', 'E7']
    gender_classes = [sheet[cell].value for cell in gender_classes_cells]

    income_class_range = 'A11:A54'
    income_classes = [cell[0].value for cell in sheet[income_class_range]]
    income_classes = [re.sub('tkr', '', str(value).rstrip().replace('–', '-'))
                      for value in income_classes]

    start_x = 11
    start_y = 3

    data = []
    for age_i, age in enumerate(age_ranges):
        for gender_i, gender in enumerate(gender_classes):
            for income_i, income in enumerate(income_classes):
                value = sheet.cell(start_x + income_i,
                                   (start_y+gender_i) + (4*age_i)).value
                value = 0 if value == '..' else value
                if income == '0':
                    min_income = 0
                    max_income = 1
                elif income == '3 000-':
                    min_income = 3000
                    max_income = None
                else:
                    min_income, max_income = [
                        int(re.sub("\D", "", x)) for x in income.split('-')]

                data.append(
                    (age, gender, income, min_income, max_income, value))

    return data


def combine_high_income_rows(df):
    min_income_threshold = 1500
    income_range_string = f'{min_income_threshold:,d}-'.replace(',', ' ')
    mask = df['min_income'] >= min_income_threshold
    group_by_headers = ['age_range', 'gender',
                        'min_income', 'income', 'max_income']
    df.loc[mask, 'min_income'] = min_income_threshold
    df.loc[mask, 'max_income'] = ''
    df.loc[mask, 'income'] = income_range_string
    df = df.groupby(group_by_headers, as_index=False).sum()
    return df


# Processing data
def calculate_percentiles(df):
    group_by_headers = ['age_range', 'gender']
    df['percent_of_people'] = df['number_of_people'] / \
        df.groupby(group_by_headers)['number_of_people'].transform('sum')
    df['percentile'] = df.groupby(group_by_headers)[
        'percent_of_people'].cumsum()
    df['percentile'] = df['percentile'] * 100
    df['selected_income_bracket'] = False
    return df


def create_income_df(df_in):
    df = df_in.copy()
    df = combine_high_income_rows(df)
    df = calculate_percentiles(df)
    return df


def create_interpolation_series(df, percentiles):
    x = df['percentile'].to_numpy()
    y = df['min_income'].to_numpy()
    return np.interp(percentiles, x, y, left=0)


def create_interpolated_percentile_df(df_in):
    df = df_in.copy()
    df = calculate_percentiles(df)
    percentile_dfs_list = []
    percentiles = list(range(0, 101))
    group_by_headers = ['age_range', 'gender']

    for headers, group in df.groupby(group_by_headers):
        interpolated_incomes = create_interpolation_series(group, percentiles)
        group_df = pd.DataFrame({
            'percentile': percentiles,
            'interpolated_income': interpolated_incomes
        })
        group_df[group_by_headers[0]] = headers[0]
        group_df[group_by_headers[1]] = headers[1]
        percentile_dfs_list.append(group_df)
    percentile_dfs = pd.concat(percentile_dfs_list)
    percentile_dfs['selected_income_bracket'] = False
    return percentile_dfs


def process_data(df):
    df_income = create_income_df(df)
    df_interp = create_interpolated_percentile_df(df)
    return df_income, df_interp


def main():
    data = process_workbook(WB_FILE)
    df = pd.DataFrame(data, columns=COLUMNS)
    df_income, df_interp = process_data(df)
    df_income.to_csv(OUT_FILE_INCOME, index=False)
    df_interp.to_csv(OUT_FILE_PERCENTILES, index=False)


if __name__ == '__main__':
    main()
